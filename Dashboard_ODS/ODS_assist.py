import pickle
import time
from tkinter.constants import ACTIVE, DISABLED
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from PIL import Image, ImageDraw
from datetime import datetime
from openpyxl import Workbook,load_workbook
import glob
import os
import tkinter as tk
from datetime import datetime
import requests
import threading
import json

# Variables
chrome_path = r'D:\Process Improvement Project\python_programming\chromedriver.exe'
url_home = 'http://10.125.64.86/ODS/Login.html'
url_ioms = 'http://10.125.64.81/ioms'
url_ioms_status = 'http://10.125.64.81/IOMS/DOMS/Status'
ioms_username = '41018328'
ioms_password = 'pass#123'

dts_table_xpath = '//*[@id="GraphicLayer_DT_layer"]'
dts_xpath = '//*[@id="GraphicLayer_DT_layer"]/image[1]'
key_xpath = '//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr[{}]/th[1]'
value_xpath = '//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr[{}]/td'
planned_xpath = '//*[@id="popover205995"]/div[2]/div/div[1]/input'
unplanned_xpath = '//*[@id="popover205995"]/div[2]/div/div[2]/input'
amr_xpath = '//*[@id="popover205995"]/div[2]/div/div[3]/input'

driver = ''
error,info,keys = [],[],[]
start_time = time.time()
ods_window = ''
ioms_window = ''
scada_window = ''
timestamp = ''
outageid = ''
grid_name = ''
feeder_name = ''
tripping_date = ''

def click_element(ele_list):
    count_pass = []
    count_fail = []
    for ele in ele_list:
        try:
            ele.click()
            read_table()
            count_pass.append(ele)            
        except Exception as e:
            error.append(e.args)
            count_fail.append(ele)
    return count_pass,count_fail

def click_by_loc(failed):
    count_pass = []
    count_fail = []
    for dt in failed:
        ele_body = driver.find_element_by_tag_name('body')
        action = ActionChains(driver)
        x = dt.location['x']
        y = dt.location['y']
        try:
            action.move_to_element_with_offset(ele_body, x, y).click().perform()
            read_table()
            count_pass.append(dt)
        except Exception as e:
            error.append(e.args)
            count_fail.append(dt)
    return count_pass,count_fail

def read_table():
    global outageid
    global grid_name
    global feeder_name
    global tripping_date
    driver.switch_to.window(ods_window)
    next = driver.find_element_by_xpath('//*[@id="ODSMap_root"]/div[3]/div[1]/div[1]/div/div[4]')
    d={}
    d['screen_shot'] = []
    while True:
        tr_count = len(driver.find_elements_by_xpath('//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr'))
        if tr_count > 0:
            for n in range(1,tr_count+1):
                key = driver.find_element_by_xpath(key_xpath.format(n)).text
                value = driver.find_element_by_xpath(value_xpath.format(n)).text
                if key in d.keys():
                   if value not in d[key]:
                       d[key].append(value)
                else:
                    d[key] = [value]
            outageid = d['Outage ID'][0]
            d['screen_shot'].append(screenshot(outageid + '_'))
            info.append(d)
        if 'hidden' in next.get_attribute('class'):
            break
        else:
            next.click()
    grid_name = d['Grid Name'][0]
    feeder_name = d['Feeder Name'][0]
    tripping_date = d['OPEN TIME'][0]
    btn_search_detail.config(state=ACTIVE)
    # btn_search_scada.config(state=ACTIVE)
    # driver.find_element_by_xpath('//*[@id="ODSMap_root"]/div[3]/div[1]/div[1]/div/div[6]').click() 

def write_excel():
    if len(info) > 0:
        for i in info:
            for k in i.keys():
                if k not in keys:
                    keys.append(k)
        wb = Workbook()
        ws = wb.active
        for c,key in enumerate(keys,start=1):
            ws.cell(1,c).value = key
        for r,d in enumerate(info,start=2):
            for c,key in enumerate(keys,start = 1):
                try:
                    ws.cell(r,c).value = ','.join('{}'.format(v) for v in d[key])
                except:
                    pass
        file_path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\excel_reports\{}.xlsx'.format(timestamp)
        wb.save(file_path)
        wb.close()

def screenshot(outageid):
    check_directory(timestamp)
    path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\screenshots\{}\{}.png'
    exsisting_files = glob.glob(path.format(timestamp,'*'))
    max_ = 0
    for file in exsisting_files:
        try:
            no =int(file.split('_')[-1][:-4])
        except:
            no = 0
        if no > max_:
            max_ = no
    filename = outageid + 'screenshot_{}'.format(max_+1)
    driver.get_screenshot_as_file(path.format(timestamp,filename))
    watermark(path.format(timestamp,filename))
    return filename

def watermark(img):
    my_img = Image.open(img)
    x,y = my_img.size
    img_edit = ImageDraw.Draw(my_img)
    text = 'Screenshot Time :\n' + str(datetime.now().strftime('%H:%M:%S - %d/%m/%Y'))
    img_edit.text((x-150,y-50),text,fill=(0,0,0))
    my_img.save(img)

def print_summary():
    pass

def login_ODS():
    driver.switch_to.window(ods_window)
    global timestamp
    driver.get(url_home)
    time.sleep(5)
    driver.find_element_by_id('EmpID').send_keys('DASHBOARD')
    driver.find_element_by_id('Password').send_keys('bses@1234')
    driver.find_element_by_xpath('//*[@id="LoginForm"]/div[3]/div[2]/button').click()
    timestamp = '{}'.format(datetime.now().strftime('%d%m%Y_%H%M'))

def login_ioms():
    driver.switch_to.window(ioms_window)
    driver.get(url_ioms)
    driver.find_element_by_name('UserID').send_keys(ioms_username)
    driver.find_element_by_name('Password').send_keys(ioms_password)
    driver.find_element_by_xpath('//*[@id="login"]/form/table/tbody/tr[3]/td[3]/button').click()

def search_ioms():
    driver.switch_to.window(ioms_window)
    driver.refresh()
    userid = driver.find_elements_by_xpath('//*[@id="TheHeader"]/div/div[2]/ul/li[2]/a')
    if len(userid) == 0:
        login_ioms()
        time.sleep(5)
    if outageid in ['AMR','SCADA']:
        driver.get(url_ioms_status)
    else:
        id_link = 'http://10.125.64.81/IOMS/DOMS/Updatefault/{}'
        driver.get(id_link.format(outageid))

def check_directory(timestamp):
    path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\screenshots\{}'.format(timestamp)
    if not os.path.exists(path):
        os.mkdir(path)

def init_driver():
    global driver
    global ods_window
    global ioms_window
    global scada_window
    driver = webdriver.Chrome(executable_path=chrome_path)
    driver.maximize_window()
    ods_window = driver.current_window_handle
    script = 'window.open("","_blank")'
    for _ in range(2):
        driver.execute_script(script)
    ioms_window = driver.window_handles[1]
    scada_window = driver.window_handles[2]
    btn_login.config(state=ACTIVE)
    btn_read_table.config(state=ACTIVE)
    btn_screenshot1.config(state=ACTIVE)
    btn_screenshot2.config(state=ACTIVE)
    btn_screenshot3.config(state=ACTIVE)
    btn_import_excel.config(state=ACTIVE)
    btn_fetch_all.config(state=ACTIVE)
    btn_quit.config(state=ACTIVE)
    login_ODS()

def quit():
    if len(info) > 0:
        write_excel()
    driver.quit()

def fetch_all():
    driver.switch_to.window(ods_window)
    dts = driver.find_element_by_xpath('//*[@id="GraphicLayer_DT_layer"]').find_elements_by_xpath('.//*')
    feeders = driver.find_element_by_xpath('//*[@id="GraphicLayer_Feeder_layer"]').find_elements_by_xpath('.//*')
    for ele_list in [feeders,dts]:
        ele_pass, ele_fail = click_element(ele_list)
        loc_pass, loc_fail = click_by_loc(ele_fail) 
    write_excel()

def search_scada():
    gid = ''
    with open(r'D:\Process Improvement Project\python_programming\Dashboard_ODS\grid_name.dat','rb') as f:
        gridlist = pickle.load(f)
    for g in gridlist:
        if g['gname'] == grid_name:
            gid = g['gid']
            break
    scada_live(gid,feeder_name)

def scada_live(gid,feedername):
    url = 'http://10.125.64.86/Scada_Live/Base/GETFeederLoadByGrid?gridId={}&Type=All'.format(gid)
    try:
        response = requests.get(url,20)
        data = json.loads(response.json())
        for d in data:
            if d['FEEDERNAME'] == feedername:
                scada_live_page(d)
                break
    except Exception as e:
        print(e)
        pass

def scada_live_page(d):
    with open(r'D:\Process Improvement Project\python_programming\Dashboard_ODS\template\scada_template.html','r') as f:
        html_temp = f.read()
    html_text = html_temp.format(d['GRIDNAME'],d['GRIDID'],d['FEEDERNAME'],d['EXTERNALSCADAID'],d['Y_SYSTIME'],
       d['EVENT'],d['O_C'],d['E_F'],d['AUTOTRIP'],d['R'],d['Y'],d['B'],d['RY'],d['BR'],d['YB'],d['Y_LSTPKLOAD_Day'],
       d['Y_LSTPKLOAD_Day_DT'],d['Y_LSTPKLOAD_Month'],d['Y_LSTPKLOAD_Month_DT'])
    with open(r'D:\Process Improvement Project\python_programming\Dashboard_ODS\template\scada_live.html','w') as f:
        f.write(html_text)
    

def get_details():
    t1 = threading.Thread(target=search_ioms,args=())
    t2 = threading.Thread(target=search_scada,args=())
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    driver.switch_to.window(scada_window)
    driver.get(r'D:\Process Improvement Project\python_programming\Dashboard_ODS\template\scada_live.html')
    
def screenshot_window(window):
    driver.switch_to.window(window)
    if window == ioms_window:
        tag = 'ioms_'
    elif window == scada_window:
        tag = 'scadalive_'
    elif window == ods_window:
        tag = 'ods_'
    screenshot(tag)

window = tk.Tk()
window.title("ODS assistant")
window.geometry('180x500')

btn_open_chrome = tk.Button(window,text='Open Browser',bd=5,command=init_driver,width=40)
btn_open_chrome.pack(padx=5,pady=5)

btn_login = tk.Button(window,text='Login ODS',bd=5,command=login_ODS,state=DISABLED,width=40)
btn_login.pack(padx=5,pady=5)

btn_read_table = tk.Button(window,text='Read Table and ScreenShot',bd=5,command=read_table,state=DISABLED,width=40)
btn_read_table.pack(padx=5,pady=5)

btn_screenshot1 = tk.Button(window,text='Screenshot-ODS',state=DISABLED,bd=5,width=40,command=lambda: screenshot_window(ods_window))
btn_screenshot1.pack(padx=5,pady=5)

btn_screenshot2 = tk.Button(window,text='Screenshot-SCADALIVE',state=DISABLED,bd=5,width=40,command=lambda: screenshot_window(scada_window))
btn_screenshot2.pack(padx=5,pady=5)

btn_screenshot3 = tk.Button(window,text='Screenshot-IOMS',state=DISABLED,bd=5,width=40,command=lambda: screenshot_window(ioms_window))
btn_screenshot3.pack(padx=5,pady=5)

btn_import_excel = tk.Button(window,text='Import Excel',bd=5,width=40,command=write_excel,state=DISABLED)
btn_import_excel.pack(padx=5,pady=5)

btn_fetch_all = tk.Button(window,text='Fetch all',bd=5,width=40,command=fetch_all,state=DISABLED)
btn_fetch_all.pack(padx=5,pady=5)

btn_search_detail = tk.Button(window,text='Get Details',bd=5,width=40,command=get_details,state=DISABLED)
btn_search_detail.pack(padx=5,pady=5)

btn_search_ioms = tk.Button(window,text='Search in IOMS',bd=5,width=40,command=search_ioms,state=DISABLED)
btn_search_ioms.pack(padx=5,pady=5)

btn_search_scada = tk.Button(window,text='SCADA Live details',bd=5,width=40,command=search_scada,state=DISABLED)
btn_search_scada.pack(padx=5,pady=5)

btn_quit = tk.Button(window,text='Finish',bd=5,width=40,command=quit,state=DISABLED)
btn_quit.pack(padx=5,pady=30)

window.attributes('-topmost',1)

window.mainloop()


        