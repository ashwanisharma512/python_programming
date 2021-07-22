import pickle
import time
from tkinter.constants import ACTIVE, DISABLED
from urllib.parse import DefragResult
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from PIL import Image, ImageDraw
from datetime import datetime
from openpyxl import Workbook,load_workbook
import glob
import os
import tkinter as tk
from datetime import datetime
import requests
import threading
import json
import win32com.client as win32

# Variables
chrome_path = r'D:\Process Improvement Project\python_programming\chromedriver.exe'
url_home = 'http://10.125.64.86/ODS/Login.html'
url_ioms = 'http://10.125.64.81/ioms'
url_ioms_status = 'http://10.125.64.81/IOMS/DOMS/Status'
ioms_username = '41018328'
ioms_password = 'pass#123'
url_ioms_report = 'http://10.125.64.81/IOMS/Reports/{}'
ioms_bd = 'breakdowntmis'
ioms_psd = 'pshutmis'
ioms_em = 'eshutmis'
ioms_intbd = 'internalmis'

dts_table_xpath = '//*[@id="GraphicLayer_DT_layer"]'
dts_xpath = '//*[@id="GraphicLayer_DT_layer"]/image[1]'
key_xpath = '//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr[{}]/th[1]'
value_xpath = '//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr[{}]/td'
planned_xpath = '//*[@id="popover205995"]/div[2]/div/div[1]/input'
unplanned_xpath = '//*[@id="popover205995"]/div[2]/div/div[2]/input'
amr_xpath = '//*[@id="popover205995"]/div[2]/div/div[3]/input'

driver1 = ''
driver2 = ''
error,info,keys = [],[],[]
ioms_window = ''
scada_window = ''
timestamp = ''
outageid = ''
grid_name = ''
feeder_name = ''
tripping_date = ''
ioms_d = ''
ods_data,scadalive_data,ioms_data,sanket_data = [],[],[],[]
excel = ''
row = 1
wb = ''
ws = ''

def init_driver():
    global driver1
    global driver2
    driver1 = webdriver.Chrome(executable_path=chrome_path)
    driver1.maximize_window()
    driver2 = webdriver.Chrome(executable_path=chrome_path)
    driver2.maximize_window()
    login_ODS()

def login_ODS():
    global timestamp
    driver1.get(url_home)
    time.sleep(5)
    driver1.find_element_by_id('EmpID').send_keys('DASHBOARD')
    driver1.find_element_by_id('Password').send_keys('bses@1234')
    driver1.find_element_by_xpath('//*[@id="LoginForm"]/div[3]/div[2]/button').click()
    timestamp = '{}'.format(datetime.now().strftime('%d%m%Y_%H%M'))

def read_table():
    ioms_d = {}
    global outageid
    global grid_name
    global feeder_name
    global tripping_date
    next = driver1.find_element_by_xpath('//*[@id="ODSMap_root"]/div[3]/div[1]/div[1]/div/div[4]')
    d={}
    d['screen_shot'] = []
    round = 1
    while True:
        tr_count = len(driver1.find_elements_by_xpath('//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr'))
        if tr_count > 0:
            for n in range(1,tr_count+1):
                key = driver1.find_element_by_xpath(key_xpath.format(n)).text
                value = driver1.find_element_by_xpath(value_xpath.format(n)).text
                if key in d.keys():
                   if value not in d[key]:
                       d[key].append(value)
                else:
                    d[key] = [value]
            d['screen_shot'].append(screenshot(outageid + '_',driver1))
        if round == 1:
            outageid = d['Outage ID'][0]
            grid_name = d['Grid Name'][0]
            feeder_name = d['Feeder Name'][0]
            tripping_date = d['OPEN TIME'][0]
            t1 = threading.Thread(target=search_ioms,args=())
            t2 = threading.Thread(target=search_scada,args=())
            t1.start()
            t2.start()
            round = 2
        if 'hidden' in next.get_attribute('class'):
            break
        else:
            next.click()
    info.append(d)
    t1.join()
    t2.join()
    report_in_excel()

def write_excel():
    if len(info) > 0:
        for i in info:
            for k in i.keys():
                if k not in keys:
                    keys.append(k)
        wb = Workbook()
        ws = wb.active
        for c,key in enumerate(keys,start=1):
            ws.cell(1,c).value = key
        for r,d in enumerate(info,start=2):
            for c,key in enumerate(keys,start = 1):
                try:
                    ws.cell(r,c).value = ','.join('{}'.format(v) for v in d[key])
                except:
                    pass
        file_path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\excel_reports\{}.xlsx'.format(timestamp)
        wb.save(file_path)
        wb.close()

def report_in_excel():
    global scadalive_data,ioms_data,excel,row,wb,ws
    scadalive_keys = ['GRIDNAME','GRIDID','FEEDERNAME','EXTERNALSCADAID','Y_SYSTIME','EVENT','O_C','E_F','AUTOTRIP','R','Y','B','RY','BR','YB','Y_LSTPKLOAD_Day','Y_LSTPKLOAD_Day_DT',
    'Y_LSTPKLOAD_Month','Y_LSTPKLOAD_Month_DT']
    if excel == '':
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Add()
        ws = wb.Worksheets.Add()
        ws.Name = 'ODS Report'
    ws.Cells(row, 1).Value = 'As per ODS'
    row +=1
    for col,key in enumerate(info[-1].keys(),start=1):
        ws.Cells(row,col).Value = key
        ws.Cells(row+1,col).Value = info[-1][key]
    row +=3
    ws.Cells(row, 1).Value = 'As per SCADALIVE'
    row +=1
    for col,key in enumerate(scadalive_keys,start=1):
        ws.Cells(row,col).Value = key
        ws.Cells(row+1,col).Value = scadalive_data[key]
    row +=3
    ws.Cells(row, 1).Value = 'As per IOMS'
    row +=1
    for col,key in enumerate(ioms_data.keys(),start=1):
        ws.Cells(row,col).Value = key
        ws.Cells(row+1,col).Value = ioms_data[key]
    row +=3

def screenshot(outageid,driver):
    check_directory(timestamp)
    path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\screenshots\{}\{}.png'
    exsisting_files = glob.glob(path.format(timestamp,'*'))
    max_ = 0
    for file in exsisting_files:
        try:
            no =int(file.split('_')[-1][:-4])
        except:
            no = 0
        if no > max_:
            max_ = no
    filename = outageid + 'screenshot_{}'.format(max_+1)
    driver.get_screenshot_as_file(path.format(timestamp,filename))
    watermark(path.format(timestamp,filename))
    return filename

def check_directory(timestamp):
    path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\screenshots\{}'.format(timestamp)
    if not os.path.exists(path):
        os.mkdir(path)

def watermark(img):
    my_img = Image.open(img)
    x,y = my_img.size
    img_edit = ImageDraw.Draw(my_img)
    text = 'Screenshot Time :\n' + str(datetime.now().strftime('%H:%M:%S - %d/%m/%Y'))
    img_edit.text((x-150,y-50),text,fill=(0,0,0))
    my_img.save(img)

def search_ioms():
    global ioms_d
    userid = driver2.find_elements_by_xpath('//*[@id="TheHeader"]/div/div[2]/ul/li[2]/a')
    if len(userid) == 0:
        driver2.get(url_ioms)
        driver2.find_element_by_name('UserID').send_keys(ioms_username)
        driver2.find_element_by_name('Password').send_keys(ioms_password)
        driver2.find_element_by_xpath('//*[@id="login"]/form/table/tbody/tr[3]/td[3]/button').click()
    if outageid in ['AMR','SCADA']:       
        # driver2.get(url_ioms_report)
        ioms_d = {'DATA STATUS' : 'NO DATA FOUND'}
    else:
        id_link = 'http://10.125.64.81/IOMS/DOMS/Updatefault/{}'
        driver2.get(id_link.format(outageid))
        if 'Status' in driver2.current_url:
            if outageid[0] == 'B':
                id_link = url_ioms_report.format(ioms_bd)
            elif outageid[0] == 'P':
                id_link = url_ioms_report.format(ioms_psd)
            elif outageid[0] == 'I':
                id_link = url_ioms_report.format(ioms_intbd)
            elif outageid[0] == 'E':
                id_link = url_ioms_report.format(ioms_em)
            driver2.get(id_link)
            script1 = '$("#bdfrom").val("{} 00:00")'.format(tripping_date.split(' ')[0])
            script2 = '$("#bdto").val("{} 23:30")'.format(datetime.today().strftime('%d-%m-%Y'))
            driver2.execute_script(script1)
            driver2.execute_script(script2)
            driver2.find_element_by_xpath('//*[@id="showbd"]').click()
            imgloader()
            table_body_row_count = len(driver2.find_elements_by_xpath('//*[@id="gridtable"]/tbody/tr'))
            rowno = -1
            for no in range(1,table_body_row_count+1):
                if driver2.find_element_by_xpath('//*[@id="gridtable"]/tbody/tr[{}]/td[3]'.format(no)).text == outageid:
                    rowno = no
                    break
            if rowno > 0:
                no_of_col = len(driver2.find_elements_by_xpath('//*[@id="gridtable"]/thead/tr/th'))
                for n in range(1,no_of_col):
                    key = driver2.find_element_by_xpath('//*[@id="gridtable"]/thead/tr/th[{}]'.format(n)).text
                    val = driver2.find_element_by_xpath('//*[@id="gridtable"]/tbody/tr[{}]/td[{}]'.format(rowno,n)).text
                    ioms_d[key] = val
            else:
                ioms_d = {'DATA STATUS' : 'NO DATA FOUND'}
        else:
            ioms_d = {'DATA STATUS' : 'NO DATA FOUND'}
    ioms_data = ioms_d

def imgloader():
    img = driver2.find_element_by_id('imageLoader')
    while True:
        if 'none' in img.get_attribute('style'):
            break

def quit():
    if len(info) > 0:
        write_excel()
    driver1.quit()
    driver2.quit()

def search_scada():
    gid = ''
    with open(r'D:\Process Improvement Project\python_programming\Dashboard_ODS\grid_name.dat','rb') as f:
        gridlist = pickle.load(f)
    for g in gridlist:
        if g['gname'] == grid_name:
            gid = g['gid']
            break
    url = 'http://10.125.64.86/Scada_Live/Base/GETFeederLoadByGrid?gridId={}&Type=All'.format(gid)
    try:
        response = requests.get(url,20)
        data = json.loads(response.json())
        for d in data:
            if d['FEEDERNAME'] == feeder_name:
                scadalive_data = d
                break
    except Exception as e:
        print(e)
    finally:
        scadalive_data = {}

# def scada_live_page(d):
    global ioms_d
    with open(r'D:\Process Improvement Project\python_programming\Dashboard_ODS\template\scada_template.html','r') as f:
        html_temp = f.read()
    while True:
        if len(ioms_d) > 0:
            break
    ioms_data = ''
    for key in ioms_d.keys():
        ioms_data += '<tr><th>{}</th><td>{}</td></tr>'.format(key,ioms_d[key])
    html_text = html_temp.format(d['GRIDNAME'],d['GRIDID'],d['FEEDERNAME'],d['EXTERNALSCADAID'],d['Y_SYSTIME'],
       d['EVENT'],d['O_C'],d['E_F'],d['AUTOTRIP'],d['R'],d['Y'],d['B'],d['RY'],d['BR'],d['YB'],d['Y_LSTPKLOAD_Day'],
       d['Y_LSTPKLOAD_Day_DT'],d['Y_LSTPKLOAD_Month'],d['Y_LSTPKLOAD_Month_DT'],ioms_data)
    with open(r'D:\Process Improvement Project\python_programming\Dashboard_ODS\template\scada_live.html','w') as f:
        f.write(html_text)
    
def screenshot_window(window):
    driver2.switch_to.window(window)
    if window == ioms_window:
        tag = 'ioms_'
    elif window == scada_window:
        tag = 'scadalive_'
    screenshot(tag,driver2)

window = tk.Tk()
window.title("ODS assistant")
window.geometry('180x500')

btn_open_chrome = tk.Button(window,text='Open Browser',bd=5,command=init_driver,width=40)
btn_open_chrome.pack(padx=5,pady=5)

btn_login = tk.Button(window,text='Login ODS',bd=5,command=login_ODS,state=DISABLED,width=40)
btn_login.pack(padx=5,pady=5)

btn_read_table = tk.Button(window,text='Read Table and ScreenShot',bd=5,command=read_table,state=DISABLED,width=40)
btn_read_table.pack(padx=5,pady=5)

btn_screenshot1 = tk.Button(window,text='Screenshot-ODS',state=DISABLED,bd=5,width=40,command=lambda: screenshot('ods_',driver1))
btn_screenshot1.pack(padx=5,pady=5)

btn_screenshot2 = tk.Button(window,text='Screenshot-SCADALIVE',state=DISABLED,bd=5,width=40,command=lambda: screenshot_window(scada_window))
btn_screenshot2.pack(padx=5,pady=5)

btn_screenshot3 = tk.Button(window,text='Screenshot-IOMS',state=DISABLED,bd=5,width=40,command=lambda: screenshot_window(ioms_window))
btn_screenshot3.pack(padx=5,pady=5)

btn_import_excel = tk.Button(window,text='Import Excel',bd=5,width=40,command=write_excel,state=DISABLED)
btn_import_excel.pack(padx=5,pady=5)

btn_quit = tk.Button(window,text='Finish',bd=5,width=40,command=quit,state=DISABLED)
btn_quit.pack(padx=5,pady=30)

window.attributes('-topmost',1)

# window.mainloop()


        