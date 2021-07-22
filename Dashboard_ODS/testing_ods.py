import time
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from PIL import Image, ImageDraw
from datetime import datetime
from openpyxl import Workbook,load_workbook
import glob
import os
import pickle
from datetime import datetime

start_time = time.time()

timestamp = ''

def click_element(ele_list):
    count_pass = []
    count_fail = []
    for ele in ele_list:
        try:
            ele.click()
            read_table()
            count_pass.append(ele)            
        except Exception as e:
            error.append(e.args)
            count_fail.append(ele)
    return count_pass,count_fail

def click_by_loc(failed):
    count_pass = []
    count_fail = []
    for dt in failed:
        ele_body = driver.find_element_by_tag_name('body')
        action = ActionChains(driver)
        x = dt.location['x']
        y = dt.location['y']
        try:
            action.move_to_element_with_offset(ele_body, x, y).click().perform()
            read_table()
            count_pass.append(dt)
        except Exception as e:
            error.append(e.args)
            count_fail.append(dt)
    return count_pass,count_fail

def read_table():
    time.sleep(0.5)
    next = driver.find_element_by_xpath('//*[@id="ODSMap_root"]/div[3]/div[1]/div[1]/div/div[4]')
    while True:
        tr_count = len(driver.find_elements_by_xpath('//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr'))
        d={}
        if tr_count > 0:
            for n in range(1,tr_count+1):
                key = driver.find_element_by_xpath(key_xpath.format(n)).text
                value = driver.find_element_by_xpath(value_xpath.format(n)).text
                d[key] = value
            if 'Outage ID' in d.keys():
                outageid = d['Outage ID'] + '_'
            else:
                outageid = '_'
            d['screen_shot'] = screenshot(outageid)
            info.append(d)
        if 'hidden' in next.get_attribute('class'):
            break
        else:
            next.click()
    driver.find_element_by_xpath('//*[@id="ODSMap_root"]/div[3]/div[1]/div[1]/div/div[6]').click()

def write_excel():
    for i in info:
        for k in i.keys():
            if k not in keys:
                keys.append(k)
    wb = Workbook()
    ws = wb.active
    for c,key in enumerate(keys,start=1):
        ws.cell(1,c).value = key
    for r,d in enumerate(info,start=2):
        for c,key in enumerate(keys,start = 1):
            try:
                ws.cell(r,c).value = d[key]
            except:
                pass
    file_path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\excel_reports\{}.xlsx'.format(timestamp)
    wb.save(file_path)
    wb.close()

def screenshot(outageid):
    path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\screenshots\{}\{}.png'
    exsisting_files = glob.glob(path.format(timestamp,'*'))
    max_ = 0
    for file in exsisting_files:
        try:
            no =int(file.split('_')[-1][:-4])
        except:
            no = 0
        if no > max_:
            max_ = no
    filename = outageid + 'screenshot_{}'.format(max_+1)
    driver.get_screenshot_as_file(path.format(timestamp,filename))
    watermark(path.format(timestamp,filename))
    return filename

def watermark(img):
    my_img = Image.open(img)
    x,y = my_img.size
    img_edit = ImageDraw.Draw(my_img)
    text = 'Screenshot Time :\n' + str(datetime.now().strftime('%H:%M:%S - %d/%m/%Y'))
    img_edit.text((x-150,y-50),text,fill=(0,0,0))
    my_img.save(img)


def print_summary():
    print('*'*40,"Summary",'*'*40,'\n')
    print('No of Element Found\t\t',len(dts))
    print('No of Information Collected\t\t',len(info))
    print('No of failed in element click\t\t',len(ele_fail))
    print('No of failed in loc click\t\t',len(loc_fail))
    print('No of keys\t\t',len(keys))
    print('Total No of errors\t\t',len(error))
    print('Time Taken \t\t',round(time.time()-start_time,3),3)
    print('*'*42,"END",'*'*42,'\n')

def login():
    driver.get(url_home)
    time.sleep(5)
    driver.find_element_by_id('EmpID').send_keys('DASHBOARD')
    driver.find_element_by_id('Password').send_keys('bses@1234')
    driver.find_element_by_xpath('//*[@id="LoginForm"]/div[3]/div[2]/button').click()

def check_directory(timestamp):
    path = r'D:\Process Improvement Project\python_programming\Dashboard_ODS\screenshots\{}'.format(timestamp)
    if not os.path.exists(path):
        os.mkdir(path)

# Variables
chrome_path = r'D:\Process Improvement Project\python_programming\chromedriver.exe'
url_home = 'http://10.125.64.86/ODS/Login.html'
url_ioms = 'http://10.125.64.81/ioms'
ioms_username = '41018328'
ioms_password = 'pass#123'

dts_table_xpath = '//*[@id="GraphicLayer_DT_layer"]'
dts_xpath = '//*[@id="GraphicLayer_DT_layer"]/image[1]'
key_xpath = '//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr[{}]/th[1]'
value_xpath = '//*[@id="ODSMap_root"]/div[3]/div[1]/div[2]/div/table/tbody/tr[{}]/td'
planned_xpath = '//*[@id="popover205995"]/div[2]/div/div[1]/input'
unplanned_xpath = '//*[@id="popover205995"]/div[2]/div/div[2]/input'
amr_xpath = '//*[@id="popover205995"]/div[2]/div/div[3]/input'
# GraphicLayer_Feeder_layer

success = False
for _ in range(5):
    try:
        driver = webdriver.Chrome(executable_path=chrome_path)
        driver.maximize_window()
        login()
        timestamp = '{}'.format(datetime.now().strftime('%d%m%Y_%H%M'))
        check_directory(timestamp)
        time.sleep(200)
        screenshot('main_')
        driver.find_element_by_xpath('/html/body/div[2]/div/section[2]/div[3]/div/div[1]/div').click()
        screenshot('wout_table_')
        error,info,keys = [],[],[]
        dts = driver.find_element_by_xpath('//*[@id="GraphicLayer_DT_layer"]').find_elements_by_xpath('.//*')
        feeders = driver.find_element_by_xpath('//*[@id="GraphicLayer_Feeder_layer"]').find_elements_by_xpath('.//*')
        for ele_list in [feeders,dts]:
            ele_pass, ele_fail = click_element(ele_list)
            loc_pass, loc_fail = click_by_loc(ele_fail) 
            print_summary()
        # driver.quit()
        write_excel()
        break
    except Exception as e:
        print(e)
        driver.quit()
        pass
        