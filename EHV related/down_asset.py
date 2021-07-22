from openpyxl import load_workbook
from glob import glob
from datetime import date, datetime, timedelta
from os import path

from openpyxl.workbook.workbook import Workbook


x_path = r'E:\mis 20-21\BRPL MIS REPORT AS ON {}.xlsx'
maxrow = 60
maxcol = 20
text1 = ['SOUTH','WEST','PUT OFF','CES/P&C/KCC/PROTECTION ISSUES']

final_list = []
not_found = []
start_date = datetime(2020, 4, 1)

for delta in range(365):
    new_date = start_date + timedelta(delta)
    dt = new_date.strftime('%d.%m.%Y')
    file_path = x_path.format(dt)
    if path.exists(file_path):
        print(dt,end='\t')
        d = {}
        wb = load_workbook(file_path)
        ws = wb.active
        for r in range(1,maxrow+1):
            for c in range(1,maxcol):
                x = ws.cell(r,c).value
                if x in text1 and c != 6:
                    d[x] = r
        if 'CES/P&C/KCC/PROTECTION ISSUES' in d.keys():
            d2 = {
                'date' : '{}'.format(dt),
                'SOUTH' : d['WEST'] - d['SOUTH'] -1,
                'WEST' : d['PUT OFF'] - d['WEST'] -1,
                'PUT OFF' : d['CES/P&C/KCC/PROTECTION ISSUES'] - d['PUT OFF'] -1
            }
        else:
            d2 = {
                'date' : '{}'.format(dt),
                'SOUTH' : d['WEST'] - d['SOUTH'] -1,
                'WEST' : d['PUT OFF'] - d['WEST'] -1,
                'PUT OFF' : 1
            }            
        for key in d2.keys():
            if d2[key] == 1:
                for y in range(1,20):
                    if ws.cell(d[key]+y,1).value == y:
                        d2[key] = y
                    else:
                        if y == 1:
                            d2[key] = 0
                        break                
    else:
        not_found.append(file_path)
        d2 = {
            'date' : '{}'.format(dt),
            'SOUTH' : d2['SOUTH'],
            'WEST' : d2['WEST'],
            'PUT OFF' : d2['PUT OFF']
        }
        wb.close()
    final_list.append(d2)
if len(not_found) > 0:
    print('Files not found')
for nf in not_found:
    print(nf)

n = 3
m = 0
t = ''
wb1 = Workbook()
ws1 = wb1.active

for l in final_list:
    
    if t != l['date'][3:5] and n != 3:
        n = 3
        m += 6
    if n == 3:
        ws1.cell(n-2,m+3).value = '{}-2020'.format(l['date'][3:5])
        ws1.cell(n-1,m+1).value = 'date'
        ws1.cell(n-1,m+2).value = 'SOUTH'
        ws1.cell(n-1,m+3).value = 'WEST'
        ws1.cell(n-1,m+4).value = 'PUT OFF'
        ws1.cell(n-1,m+5).value = 'Total'
    ws1.cell(n,m+1).value = l['date']
    ws1.cell(n,m+2).value = l['SOUTH']
    ws1.cell(n,m+3).value = l['WEST']
    ws1.cell(n,m+4).value = l['PUT OFF']
    ws1.cell(n,m+5).value = '=SUM(A{0}:D{0})'.format(n)
    n +=1
    t = l['date'][3:5]
        

wb1.save('temp4.xlsx')
wb1.close()


n = 3
m = 0
wb1 = Workbook()
ws1 = wb1.active

for l in final_list:
    
    if n == 3:
        ws1.cell(n-1,m+1).value = 'date'
        ws1.cell(n-1,m+2).value = 'SOUTH'
        ws1.cell(n-1,m+3).value = 'WEST'
        ws1.cell(n-1,m+4).value = 'PUT OFF'
        ws1.cell(n-1,m+5).value = 'Total'
    ws1.cell(n,m+1).value = l['date']
    ws1.cell(n,m+2).value = l['SOUTH']
    ws1.cell(n,m+3).value = l['WEST']
    ws1.cell(n,m+4).value = l['PUT OFF']
    ws1.cell(n,m+5).value = '=SUM(A{0}:D{0})'.format(n)
    n +=1
        

wb1.save('temp5.xlsx')
wb1.close()