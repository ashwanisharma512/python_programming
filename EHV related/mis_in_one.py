
import glob
from openpyxl import load_workbook
from datetime import date
import re

folder_path = r'D:\Reports\Dash Board Report\2021\2021 04 Apr'

files = glob.glob(r'D:\Reports\Dash Board Report\2021\05 May 2021\daily tripping details\*Tripping details*.xlsx')

mainfile = 'MIS_template.xlsx'

wb_main = load_workbook(r'D:\Process Improvement Project\python_programming\EHV related\template\MIS_template.xlsx')
ws_main = wb_main['main']
row_main = ws_main.max_row
map_main = [3,4,5,6,7,8,9,10,11,12,15,13,14]
map_file = [2,3,4,5,6,7,8,9,10,11,12,16,17]
x = 1
trip_count = []
for file in last_7trips[-6:]:
    wb = load_workbook(file)
    ws = wb['EHV TRIPPING ']
    row_file = ws.max_row
    dt_ = re.findall("\d\d.\d\d.\d\d\d\d",ws.cell(row=1,column=1).value)
    if len(dt_) > 0:
        dt_y = int(dt_[0][-4:])
        dt_m = int(dt_[0][3:5])
        dt_d = int(dt_[0][:2])
    dat_ = date(dt_y,dt_m,dt_d)
    dat = dat_.strftime('%d-%b-%y')
    print(dat)
    tpl = (dat,row_file-2)
    trip_count.append(tpl)
    for n in range(3,row_file+1):
        ws_main.cell(row=row_main+x,column=2).value = dat
        for a,b in zip(map_main,map_file):
            ws_main.cell(row=row_main+x,column=a).value = ws.cell(row=n,column=b).value
        x += 1
    wb.close()
tp = wb_main['trip_count']

for n,count in enumerate(trip_count,start=2):
    tp.cell(row=n,column=1).value = count[0]
    tp.cell(row=n,column=2).value = count[1]
tp.cell(row=len(trip_count)+2,column=1).value = 'Total'
tp.cell(row=len(trip_count)+2,column=2).value = '=sum(B2:B{})'.format(len(trip_count)+1)

wb_main.save('MIS_complete.xlsx')
wb_main.close()